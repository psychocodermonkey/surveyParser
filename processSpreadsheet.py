#! /usr/bin/env python3
"""
# Program: Process survey results to build list of mentions correlated to comments.
#    Name: Andrew Dixon            File: processSpreadsheet .py
#    Date: 3 Mar 2026
#   Notes: Excel file has specific format and column order. See README for guidance.
#
#........1.........2.........3.........4.........5.........6.........7.........8.........9.........0.........1.........2.........3..
"""

import sqlite3
from dataclasses import dataclass
from pathlib import Path
from openpyxl import Workbook, load_workbook

# Settings (future use)
CALC_LENGTHS = True

# Declare input and output files (specify paths as necessary)
DATA_DIR: Path =  Path(__file__).parent / "data"
EXCEL_FILE: Path = DATA_DIR / "import.xlsx"
OUTPUT_FILE: Path = DATA_DIR / "name_mentions.xlsx"
SQLITE_FILE: Path = DATA_DIR / "name_mentions.sqlite"
EXCLUDE_FILE: Path = DATA_DIR / "exclude_names.txt"
ENABLE_SQLITE = True


@dataclass
class XlatedRow:
  """Data class for spreadsheet columns, gives meaningful names in program.
  Define names for columns so that the object has names that are easy to use."""

  college: str
  department: str
  semister: str
  question: str
  answer: str
  mentions: str

  def __repr__(self) -> str:
    return f'<XlatedRow(Code({self.semister}):Item({self.department}))>'

  def __str__(self) -> str:
    return f'{self.semister}, {self.college} - ({self.department})'


def main() -> None:
  """Main"""

  # Print where the input file is
  print(f'[+] Reading: {EXCEL_FILE}')

  # Functions to convert column letters to numbers and vice-versa. Using lambda because I can.
  colNum = lambda a: 0 if a == '' else 1 + ord(a[-1]) - ord('A') + 26 * colNum(a[:-1])  # noqa: E731
  colName = lambda n: '' if n <= 0 else colName((n - 1) // 26) + chr((n - 1) % 26 + ord('A'))  # noqa: E731

  # Open the defined Excel document.
  wb = load_workbook(EXCEL_FILE)
  ws = wb.active

  # Create dictionary based on the dataclass object to store the max length of each column.
  excel_data = []

  # Process the rows of the spreadsheet and put them in the dataclass object.
  for row in ws.iter_rows(values_only=True, min_row=2):
    # Map each column to it's designated propert in the dataclass
    data = XlatedRow(
      college=row[colNum('A') - 1],
      department=row[colNum('B') - 1],
      semister=row[colNum('C') - 1],
      question=row[colNum('D') - 1],
      answer=row[colNum('E') - 1],
      mentions=row[colNum('F') - 1],
    )

    # Add any clean-up or manipulation that needs to be done to data here.

    # Add the line to the data to be written out.
    excel_data.append(data)

  # Report how many rows were processed
  print(f"[+] {len(excel_data)} Rows processed.")

  excludeNames: set[str] = loadExcludeNames(EXCLUDE_FILE)
  flatRows: list[str] = []
  mentionHits = 0
  excludedMentionHits = 0
  seenNames: set[str] = set()

  conn = None
  if ENABLE_SQLITE:
    conn: sqlite3.Connection = initDatabase(SQLITE_FILE)

  for data in excel_data:
    parsedMentions: list[str] = parseMentions(data.mentions)
    filteredMentions: list[str] = []

    for mention in parsedMentions:
      normalized: str = mention.lower()

      if normalized in excludeNames:
        excludedMentionHits += 1
        continue

      filteredMentions.append(mention)
      seenNames.add(normalized)
      mentionHits += 1
      flatRows.append((mention, data.college, data.department, data.answer))

    if conn is not None:
      questionId: int = getOrCreateQuestion(conn, data.question)
      responseId: int = insertResponse(
        conn=conn,
        college=data.college,
        department=data.department,
        semister=data.semister,
        answer=data.answer,
        questionId=questionId,
      )

      # if excludedMentionHits + mentionHits < 50:
      #   print(f"[DBG] responseId={responseId} question={data.question!r}")
      #   print(f"[DBG] answer={data.answer!r}")
      #   print(f"[DBG] mentionsRaw={data.mentions!r}")
      #   print(f"[DBG] mentionsParsed={filteredMentions!r}")

      for mention in filteredMentions:
        nameId: int = getOrCreateName(conn, mention)
        conn.execute(
          "INSERT OR IGNORE INTO response_names (responseId, nameId) VALUES (?, ?)",
          (responseId, nameId),
        )

  if conn is not None:
    conn.commit()
    conn.close()

  writeOutputWorkbook(OUTPUT_FILE, flatRows)

  print(f"Rows processed: {len(excel_data)}")
  print(f"Unique names: {len(seenNames)}")
  print(f"Mention hits: {mentionHits}")
  print(f"Excluded names encountered: {excludedMentionHits}")
  print(f"Spreadsheet output: {OUTPUT_FILE}")
  print(f"SQLite database: {SQLITE_FILE}")


def loadExcludeNames(excludeFile: str) -> set[str]:
  """Load excluded names as a normalized lowercase set."""

  path = Path(excludeFile)
  if not path.exists():
    return set()

  excluded: set[str] = set()

  with path.open("r", encoding="utf-8") as handle:
    for rawLine in handle:
      line: str = rawLine.strip()

      if not line or line.startswith("#"):
        continue

      excluded.add(line.lower())

  return excluded


def parseMentions(mentions: str | None) -> list[str]:
  """Parse semicolon-delimited mentions with case-insensitive de-duplication."""

  if mentions is None:
    return []

  rawText: str = str(mentions).strip()
  if rawText == "":
    return []

  parsed: list[str] = []
  seen: set[str] = set()

  for part in rawText.split(";"):
    name: str = part.strip()

    # Skip empty names
    if not name:
      continue

    normalized: str = name.lower()

    # Skip names already seen
    if normalized in seen:
      continue

    seen.add(normalized)
    parsed.append(name)

  return parsed


def initDatabase(sqliteFile: str) -> sqlite3.Connection:
  """Initialize normalized SQLite schema."""

  # Build the database
  dbPath = Path(sqliteFile)
  if dbPath.exists():
    dbPath.unlink()

  # Build database connection to be returned
  conn: sqlite3.Connection = sqlite3.connect(sqliteFile)

  # Create the names table
  conn.execute(
    """
    CREATE TABLE names (
      nameId INTEGER PRIMARY KEY,
      displayName TEXT,
      normalizedName TEXT UNIQUE
    )
    """
  )

  # Create question table
  conn.execute(
    """
    CREATE TABLE questions (
      questionId INTEGER PRIMARY KEY,
      questionText TEXT UNIQUE
    )
    """
  )

  # Create the responses table
  conn.execute(
    """
    CREATE TABLE responses (
      responseId INTEGER PRIMARY KEY,
      college TEXT,
      department TEXT,
      semister TEXT,
      answer TEXT,
      questionId INTEGER
    )
    """
  )

  # Create the name/responses junction table
  conn.execute(
    """
    CREATE TABLE response_names (
      responseId INTEGER,
      nameId INTEGER,
      PRIMARY KEY (responseId, nameId)
    )
    """
  )

  return conn


def getOrCreateName(conn: sqlite3.Connection, displayName: str) -> int:
  """Get existing or create new name record."""

  normalized: str = displayName.lower()

  row = conn.execute(
    "SELECT nameId FROM names WHERE normalizedName = ?",
    (normalized,),
  ).fetchone()

  # Return the rowid if we got one
  if row is not None:
    return int(row[0])

  # Add the row if we didn't get a hit
  cursor: sqlite3.Cursor = conn.execute(
    "INSERT INTO names (displayName, normalizedName) VALUES (?, ?)",
    (displayName, normalized),
  )

  return int(cursor.lastrowid)


def getOrCreateQuestion(conn: sqlite3.Connection, questionText: str | None) -> int:
  """Get existing or create new question record."""

  questionValue: str = "" if questionText is None else str(questionText)

  row = conn.execute(
    "SELECT questionId FROM questions WHERE questionText = ?",
    (questionValue,),
  ).fetchone()

  # Return the rowid if we got a hit
  if row is not None:
    return int(row[0])

  # Insert if we didn't get a hit
  cursor = conn.execute(
    "INSERT INTO questions (questionText) VALUES (?)",
    (questionValue,),
  )

  return int(cursor.lastrowid)


def insertResponse(
  conn: sqlite3.Connection,
  college: str | None,
  department: str | None,
  semister: str | None,
  answer: str | None,
  questionId: int,
) -> int:
  """Insert one response row and return responseId."""

  cursor = conn.execute(
    """
    INSERT INTO responses (college, department, semister, answer, questionId)
    VALUES (?, ?, ?, ?, ?)
    """,
    (
      "" if college is None else str(college),
      "" if department is None else str(department),
      "" if semister is None else str(semister),
      "" if answer is None else str(answer),
      questionId,
    ),
  )

  return int(cursor.lastrowid)


def writeOutputWorkbook(outputFile: str, rows: list[tuple[str, str, str, str]]) -> None:
  """Write flattened mention output workbook."""

  workbook = Workbook()
  sheet = workbook.active
  sheet.title = "Mentions"
  sheet.append(["Name", "College", "Department", "Answer"])

  sortedRows = sorted(
    rows,
    key=lambda item: (
      item[0].lower(),
      str(item[1]).lower(),
      str(item[2]).lower(),
    ),
  )

  for name, college, department, answer in sortedRows:
    sheet.append(
      [
        "" if name is None else str(name),
        "" if college is None else str(college),
        "" if department is None else str(department),
        "" if answer is None else str(answer),
      ]
    )

  workbook.save(outputFile)


# If the ExcelToFlat.py is run (instead of imported as a module),
# call the main() function:
if __name__ == '__main__':
  main()
